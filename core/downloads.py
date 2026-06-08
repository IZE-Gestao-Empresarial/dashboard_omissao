from __future__ import annotations

from datetime import datetime
from csv import DictReader
from io import BytesIO
from io import StringIO
from pathlib import Path
from typing import Any, Dict, Sequence
import unicodedata
from urllib.parse import quote, urlencode

import requests
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


_BASE_DIR = Path(__file__).resolve().parent.parent
_DEFAULT_LOCAL_WORKBOOK = _BASE_DIR / "data" / "abas_detalhamento_preenchida_ficticia.xlsx"
_DEFAULT_DASHBOARD_SPREADSHEET_ID = "1XKzSX1-xspp8NAOF7WGli0dwJ2SRlOdfm0fxOCUDVHs"
_DEFAULT_DETAILS_SPREADSHEET_ID = "1lOVjHVNgofpmJeEnt3zwsCWve5MJX5HO3PqbOAPmHf0"
_DEFAULT_TIMEOUT = 30
_CACHE_TTL_SECONDS = 300

_DOWNLOAD_SPECS = {
    "base": {
        "sheet_names": ["BASE_INDICADORES"],
        "filename_prefix": "base_indicadores",
        "label": "Base",
        "source": "dashboard_direct_sheets",
        "extra_sheets": [
            {
                "sheet_names": ["DETALHAMENTO_CANCELADOS"],
                "source": "direct_sheets",
            }
        ],
    },
    "atualizacao": {
        "sheet_names": ["DETALHAMENTO_ATUALIZACAO"],
        "filename_prefix": "detalhamento_atualizacao",
        "label": "Atualização",
        "source": "direct_sheets",
    },
    "rotina": {
        "sheet_names": ["DETALHAMENTO_CHURN"],
        "filename_prefix": "detalhamento_churn",
        "label": "Rotina",
        "source": "direct_sheets",
    },
    "avancado": {
        "sheet_names": ["DETALHAMENTO_AVANCADO", "DETALHAMENTO_AVANÇADO"],
        "filename_prefix": "detalhamento_avancado",
        "label": "Avançado",
        "source": "direct_sheets",
    },
    "entrega_final": {
        "sheet_names": ["DETALHAMENTO_ENTREGAS", "DETALHAMENTO_ENTREGA"],
        "filename_prefix": "detalhamento_entregas",
        "label": "Entregas Finais",
        "source": "direct_sheets",
    },
    "produto": {
        "sheet_names": ["DETALHAMENTO_PRODUTO", "DETALHAMENTO_PRODUTOS"],
        "filename_prefix": "detalhamento_produto",
        "label": "Produto",
        "source": "direct_sheets",
    },
}


def _get_secret_str(key: str, default: str = "") -> str:
    try:
        value = st.secrets.get(key, default)
    except Exception:
        return default
    return str(value).strip() if value is not None else default


def _normalize_header_name(value: str | None) -> str:
    raw = str(value or "").strip()
    if not raw:
        return "Coluna"
    raw = raw.replace("_", " ").replace("-", " ")
    parts = [part for part in raw.split() if part]
    if not parts:
        return "Coluna"

    upper_tokens = {"FC", "DRE", "ID", "CNPJ", "CPF", "API", "URL", "ETAPA"}
    titled = []
    for part in parts:
        clean = part.upper()
        if clean in upper_tokens:
            titled.append(clean)
        else:
            titled.append(part.capitalize())
    return " ".join(titled)


def _normalize_record_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Sim" if value else "Não"
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    if isinstance(value, str):
        return value.strip()
    return value


def _prettify_records(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    pretty_records: list[dict[str, Any]] = []
    for record in records:
        pretty_record: dict[str, Any] = {}
        for key, value in record.items():
            pretty_record[_normalize_header_name(str(key))] = _normalize_record_value(value)
        pretty_records.append(pretty_record)
    return pretty_records


def _strip_accents(value: str) -> str:
    return unicodedata.normalize("NFD", value).encode("ascii", "ignore").decode("ascii")


def _normalize_string_flag(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""

    lowered = _strip_accents(raw).lower()
    if lowered == "sim":
        return "Sim"
    if lowered == "nao":
        return "Não"
    if lowered == "ativo":
        return "Ativo"
    if lowered == "inativo":
        return "Inativo"
    return raw


def _filter_and_normalize_base_records(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    filtered_records: list[dict[str, Any]] = []

    for record in records:
        status_key = next((key for key in record.keys() if _strip_accents(str(key)).strip().lower() == "status"), None)
        status_value = _normalize_string_flag(record.get(status_key)) if status_key else ""
        if status_value == "Ativo":
            platform_value = "TEC"
        elif not status_value:
            platform_value = "MONDAY"
        else:
            platform_value = "Fora da Plataforma"

        normalized_record: dict[str, Any] = {}
        for key, value in record.items():
            if key == status_key:
                normalized_record["Plataforma"] = platform_value
                continue
            if isinstance(value, str):
                normalized_record[key] = _normalize_string_flag(value)
            else:
                normalized_record[key] = value

        filtered_records.append(normalized_record)

    return filtered_records


def _filter_records_by_area(records: Sequence[dict[str, Any]], area: str | None) -> list[dict[str, Any]]:
    if not area:
        return list(records)

    area_norm = _strip_accents(area).strip().upper()
    area_keys = ("Time", "TIME", "Área", "ÁREA", "Area", "AREA")
    return [
        record for record in records
        if _strip_accents(str(next((record.get(key) for key in area_keys if key in record), ""))).strip().upper() == area_norm
    ]


def _postprocess_sheet_records(sheet_name: str, records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    if sheet_name == "BASE_INDICADORES":
        return _filter_and_normalize_base_records(records)
    return list(records)


def _normalize_filename_chunk(value: str | None) -> str:
    raw = (value or "").strip()
    if not raw:
        return datetime.now().strftime("%Y-%m-%d_%H-%M")
    cleaned = []
    for char in raw:
        if char.isalnum():
            cleaned.append(char)
        elif char in {"-", "_"}:
            cleaned.append(char)
        else:
            cleaned.append("_")
    normalized = "".join(cleaned).strip("_")
    return normalized or datetime.now().strftime("%Y-%m-%d_%H-%M")


def _sheet_csv_url(spreadsheet_id: str, sheet_name: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/gviz/tq?tqx=out:csv&sheet={quote(sheet_name)}"


def _sheet_values_api_url(spreadsheet_id: str, sheet_name: str, api_key: str) -> str:
    range_name = f"'{sheet_name}'"
    params = urlencode(
        {
            "key": api_key,
            "majorDimension": "ROWS",
            "valueRenderOption": "UNFORMATTED_VALUE",
            "dateTimeRenderOption": "FORMATTED_STRING",
        }
    )
    return f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{quote(range_name, safe='')}?{params}"


def _records_from_values(values: list[list[Any]]) -> list[dict[str, Any]]:
    if not values:
        return []

    headers = [str(value).strip() for value in values[0]]
    records: list[dict[str, Any]] = []
    for value_row in values[1:]:
        record = {
            headers[index]: value_row[index] if index < len(value_row) else ""
            for index in range(len(headers))
            if headers[index]
        }
        if any(str(value).strip() for value in record.values()):
            records.append(record)
    return records


@st.cache_data(ttl=_CACHE_TTL_SECONDS, show_spinner=False)
def _fetch_direct_sheet_records(spreadsheet_id: str, sheet_name: str) -> list[dict[str, Any]]:
    api_key = _get_secret_str("GOOGLE_SHEETS_API_KEY")
    if api_key:
        response = requests.get(_sheet_values_api_url(spreadsheet_id, sheet_name, api_key), timeout=_DEFAULT_TIMEOUT)
        response.raise_for_status()
        values = response.json().get("values")
        return _records_from_values(values if isinstance(values, list) else [])

    response = requests.get(_sheet_csv_url(spreadsheet_id, sheet_name), timeout=_DEFAULT_TIMEOUT)
    response.raise_for_status()
    text = response.text.lstrip("\ufeff")
    return [dict(row) for row in DictReader(StringIO(text))]


@st.cache_data(ttl=_CACHE_TTL_SECONDS, show_spinner=False)
def _read_local_workbook(path_str: str) -> bytes:
    return Path(path_str).read_bytes()


# FIX: area adicionado e propagado corretamente para o requests.get
@st.cache_data(ttl=_CACHE_TTL_SECONDS, show_spinner=False)
def _fetch_detail_api_payload(url: str, api_key: str, sheet_name: str, area: str | None = None) -> dict[str, Any]:
    params = {"api_key": api_key, "sheet_name": sheet_name}
    if area:
        params["area"] = area
    response = requests.get(
        url,
        params=params,
        timeout=_DEFAULT_TIMEOUT,
        headers={"Cache-Control": "no-cache"},
    )
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, list):
        return {"ok": True, "rows": payload}
    if not isinstance(payload, dict):
        raise ValueError("A API de detalhamento retornou um formato inválido.")
    return payload


def _fetch_dashboard_sheet_payload(url: str, token: str, sheet_name: str, area: str | None = None) -> dict[str, Any]:
    params = {"token": token, "sheet": sheet_name}
    response = requests.get(
        url,
        params=params,
        timeout=_DEFAULT_TIMEOUT,
        headers={"Cache-Control": "no-cache"},
    )
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, list):
        return {"ok": True, "rows": payload}
    if not isinstance(payload, dict):
        raise ValueError("A API do dashboard retornou um formato inválido.")
    return payload


def _coerce_records_from_rows(rows: Sequence[Any], headers: Sequence[str] | None = None) -> list[dict[str, Any]]:
    if not rows:
        return []

    first = rows[0]
    if isinstance(first, dict):
        ordered_keys: list[str] = []
        if headers:
            ordered_keys.extend(str(col) for col in headers)
        for row in rows:
            for key in row.keys():
                key_str = str(key)
                if key_str not in ordered_keys:
                    ordered_keys.append(key_str)
        return [{key: row.get(key) for key in ordered_keys} for row in rows if isinstance(row, dict)]

    if headers:
        header_list = [str(col) for col in headers]
        normalized_records: list[dict[str, Any]] = []
        for row in rows:
            if isinstance(row, (list, tuple)):
                normalized_records.append({header_list[idx]: row[idx] if idx < len(row) else None for idx in range(len(header_list))})
        return normalized_records

    if isinstance(first, (list, tuple)) and first:
        header_list = [str(col) for col in first]
        normalized_records = []
        for row in rows[1:]:
            if isinstance(row, (list, tuple)):
                normalized_records.append({header_list[idx]: row[idx] if idx < len(row) else None for idx in range(len(header_list))})
        return normalized_records

    return []


def _extract_records(payload: dict[str, Any]) -> list[dict[str, Any]]:
    if payload.get("ok") is False:
        message = payload.get("message") or payload.get("error") or "A API de detalhamento retornou ok=false."
        raise ValueError(str(message))

    for key in ("rows", "data", "items", "records"):
        value = payload.get(key)
        if isinstance(value, list):
            headers = payload.get("headers") or payload.get("columns")
            if isinstance(headers, list):
                return _coerce_records_from_rows(value, headers=headers)
            return _coerce_records_from_rows(value)

    values = payload.get("values")
    headers = payload.get("headers") or payload.get("columns")
    if isinstance(values, list):
        return _coerce_records_from_rows(values, headers=headers if isinstance(headers, list) else None)

    raise ValueError("A API de detalhamento não retornou 'rows', 'data', 'items', 'records' ou 'values'.")


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 50)


def _records_to_xlsx_bytes(records: Sequence[dict[str, Any]], *, sheet_title: str) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = (sheet_title or "Detalhamento")[:31]
    records = _prettify_records(records)

    if not records:
        worksheet["A1"] = "Sem dados"
        worksheet["A1"].font = Font(bold=True)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    headers = list(records[0].keys())
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for record in records:
        worksheet.append([record.get(header) for header in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_columns(worksheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_records_to_worksheet(worksheet, records: Sequence[dict[str, Any]]) -> None:
    records = _prettify_records(records)
    if not records:
        worksheet["A1"] = "Sem dados"
        worksheet["A1"].font = Font(bold=True)
        return
    headers = list(records[0].keys())
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for record in records:
        worksheet.append([record.get(header) for header in headers])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_columns(worksheet)


def _records_to_xlsx_bytes_multi(sheets: list[tuple[str, list[dict[str, Any]]]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_title, records in sheets:
        worksheet = workbook.create_sheet(title=sheet_title[:31])
        _write_records_to_worksheet(worksheet, records)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _resolve_workbook_bytes() -> bytes:
    local_path = _get_secret_str("DETAIL_WORKBOOK_PATH", "")
    candidate = Path(local_path) if local_path else _DEFAULT_LOCAL_WORKBOOK
    if not candidate.exists():
        raise FileNotFoundError(
            "Arquivo de detalhamento não encontrado. Configure DETAILS_WEBAPP_URL/API_KEY ou DETAIL_WORKBOOK_PATH."
        )
    return _read_local_workbook(str(candidate))


def _copy_sheet_only(source_bytes: bytes, sheet_name: str) -> bytes:
    source_wb = load_workbook(BytesIO(source_bytes), data_only=True)
    if sheet_name not in source_wb.sheetnames:
        raise KeyError(f"Aba '{sheet_name}' não encontrada no workbook de detalhamento.")

    source_ws = source_wb[sheet_name]
    records: list[dict[str, Any]] = []
    rows = list(source_ws.iter_rows(values_only=True))
    if rows:
        headers = [str(value) if value is not None else f"COLUNA_{idx + 1}" for idx, value in enumerate(rows[0])]
        for row in rows[1:]:
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                records.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))})
    return _records_to_xlsx_bytes(records, sheet_title=sheet_name)


# FIX: area adicionado e propagado para as funções de fetch
def _fetch_xlsx_bytes_for_sheet(sheet_name: str, *, source: str = "details_api", area: str | None = None) -> bytes:
    if source == "direct_sheets":
        spreadsheet_id = _get_secret_str("DETAILS_SPREADSHEET_ID", _DEFAULT_DETAILS_SPREADSHEET_ID)
        records = _fetch_direct_sheet_records(spreadsheet_id, sheet_name)
        if area:
            area_norm = _strip_accents(area).strip().upper()
            records = [
                r for r in records
                if _strip_accents(str(r.get("Time") or r.get("TIME") or r.get("Área") or r.get("AREA") or "")).strip().upper() == area_norm
            ]
        return _records_to_xlsx_bytes(records, sheet_title=sheet_name)

    if source == "dashboard_direct_sheets":
        spreadsheet_id = _get_secret_str("DASHBOARD_SPREADSHEET_ID", _DEFAULT_DASHBOARD_SPREADSHEET_ID)
        records = _fetch_direct_sheet_records(spreadsheet_id, sheet_name)
        records = _filter_records_by_area(records, area)
        records = _postprocess_sheet_records(sheet_name, records)
        return _records_to_xlsx_bytes(records, sheet_title=sheet_name)

    if source == "dashboard_api":
        dashboard_url = _get_secret_str("SHEETS_WEBAPP_URL")
        dashboard_token = _get_secret_str("SHEETS_WEBAPP_TOKEN")
        if not dashboard_url or not dashboard_token:
            raise RuntimeError("Configure SHEETS_WEBAPP_URL e SHEETS_WEBAPP_TOKEN para baixar a aba BASE_INDICADORES.")
        payload = _fetch_dashboard_sheet_payload(dashboard_url, dashboard_token, sheet_name, area=area)
        records = _extract_records(payload)
        if area:
            area_norm = area.strip().upper()
            records = [
                r for r in records
                if str(r.get("TIME") or r.get("ÁREA") or r.get("AREA") or "").strip().upper() == area_norm
            ]
        if sheet_name == "BASE_INDICADORES":
            records = _filter_and_normalize_base_records(records)
        return _records_to_xlsx_bytes(records, sheet_title=sheet_name)

    detail_url = _get_secret_str("DETAILS_WEBAPP_URL")
    detail_api_key = _get_secret_str("DETAILS_WEBAPP_API_KEY")

    if detail_url and detail_api_key:
        payload = _fetch_detail_api_payload(detail_url, detail_api_key, sheet_name, area=area)
        records = _extract_records(payload)
        return _records_to_xlsx_bytes(records, sheet_title=sheet_name)

    workbook_bytes = _resolve_workbook_bytes()
    return _copy_sheet_only(workbook_bytes, sheet_name)


def _fetch_records_for_sheet(sheet_name: str, *, source: str = "details_api", area: str | None = None) -> list[dict[str, Any]]:
    if source == "direct_sheets":
        spreadsheet_id = _get_secret_str("DETAILS_SPREADSHEET_ID", _DEFAULT_DETAILS_SPREADSHEET_ID)
        records = _fetch_direct_sheet_records(spreadsheet_id, sheet_name)
        if area:
            area_norm = _strip_accents(area).strip().upper()
            records = [
                r for r in records
                if _strip_accents(str(r.get("Time") or r.get("TIME") or r.get("Área") or r.get("AREA") or "")).strip().upper() == area_norm
            ]
        return records

    if source == "dashboard_direct_sheets":
        spreadsheet_id = _get_secret_str("DASHBOARD_SPREADSHEET_ID", _DEFAULT_DASHBOARD_SPREADSHEET_ID)
        records = _fetch_direct_sheet_records(spreadsheet_id, sheet_name)
        records = _filter_records_by_area(records, area)
        return _postprocess_sheet_records(sheet_name, records)

    if source == "dashboard_api":
        dashboard_url = _get_secret_str("SHEETS_WEBAPP_URL")
        dashboard_token = _get_secret_str("SHEETS_WEBAPP_TOKEN")
        if not dashboard_url or not dashboard_token:
            raise RuntimeError("Configure SHEETS_WEBAPP_URL e SHEETS_WEBAPP_TOKEN.")
        payload = _fetch_dashboard_sheet_payload(dashboard_url, dashboard_token, sheet_name, area=area)
        records = _extract_records(payload)
        if area:
            area_norm = area.strip().upper()
            records = [
                r for r in records
                if str(r.get("TIME") or r.get("ÁREA") or r.get("AREA") or "").strip().upper() == area_norm
            ]
        if sheet_name == "BASE_INDICADORES":
            records = _filter_and_normalize_base_records(records)
        return records

    detail_url = _get_secret_str("DETAILS_WEBAPP_URL")
    detail_api_key = _get_secret_str("DETAILS_WEBAPP_API_KEY")
    if detail_url and detail_api_key:
        payload = _fetch_detail_api_payload(detail_url, detail_api_key, sheet_name, area=area)
        return _extract_records(payload)

    workbook_bytes = _resolve_workbook_bytes()
    source_wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
    if sheet_name not in source_wb.sheetnames:
        raise KeyError(f"Aba '{sheet_name}' não encontrada no workbook local.")
    source_ws = source_wb[sheet_name]
    rows = list(source_ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(v) if v is not None else f"COLUNA_{i + 1}" for i, v in enumerate(rows[0])]
    return [
        {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        for row in rows[1:]
        if any(c is not None and str(c).strip() != "" for c in row)
    ]


def get_download_spec(download_key: str, updated_at: str | None = None) -> dict[str, str]:
    spec = _DOWNLOAD_SPECS[download_key]
    suffix = _normalize_filename_chunk(updated_at)
    return {
        "filename": f"{spec['filename_prefix']}_{suffix}.xlsx",
        "label": spec["label"],
    }


def _build_unavailable_download_bytes(download_key: str, error: Exception) -> bytes:
    spec = _DOWNLOAD_SPECS.get(download_key, {})
    label = str(spec.get("label") or download_key)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Indisponivel"
    worksheet["A1"] = "Detalhamento indisponível"
    worksheet["A1"].font = Font(bold=True)
    worksheet["A2"] = f"Não foi possível carregar o detalhamento de {label}."
    worksheet["A3"] = str(error)
    _autosize_columns(worksheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_download_bytes(download_key: str, area: str | None = None) -> bytes:
    spec = _DOWNLOAD_SPECS[download_key]
    source = str(spec.get("source") or "details_api")
    extra_sheets_specs: list[dict[str, Any]] = spec.get("extra_sheets", [])

    if not extra_sheets_specs:
        last_error = None
        for sheet_name in spec["sheet_names"]:
            try:
                return _fetch_xlsx_bytes_for_sheet(sheet_name, source=source, area=area)
            except Exception as exc:
                last_error = exc
        raise RuntimeError(str(last_error) if last_error else "Detalhamento indisponível.")

    sheets_data: list[tuple[str, list[dict[str, Any]]]] = []

    last_error = None
    primary_records = None
    primary_sheet_name = spec["sheet_names"][0]
    for sheet_name in spec["sheet_names"]:
        try:
            primary_records = _fetch_records_for_sheet(sheet_name, source=source, area=area)
            primary_sheet_name = sheet_name
            break
        except Exception as exc:
            last_error = exc

    if primary_records is None:
        raise RuntimeError(str(last_error) if last_error else "Detalhamento indisponível.")

    sheets_data.append((primary_sheet_name, primary_records))

    for extra_spec in extra_sheets_specs:
        extra_source = str(extra_spec.get("source") or "details_api")
        extra_sheet_name = extra_spec["sheet_names"][0]
        for sheet_name in extra_spec["sheet_names"]:
            try:
                extra_records = _fetch_records_for_sheet(sheet_name, source=extra_source, area=area)
                extra_sheet_name = sheet_name
                sheets_data.append((extra_sheet_name, extra_records))
                break
            except Exception as exc:
                last_error = exc
                sheets_data.append((extra_sheet_name, []))

    return _records_to_xlsx_bytes_multi(sheets_data)


@st.cache_data(ttl=_CACHE_TTL_SECONDS, show_spinner=False)
def get_download_bytes(download_key: str, area: str | None = None) -> bytes:
    try:
        return _build_download_bytes(download_key, area=area)
    except Exception as exc:
        return _build_unavailable_download_bytes(download_key, exc)


def make_download_callable(download_key: str):
    def _builder() -> bytes:
        return get_download_bytes(download_key)

    return _builder
